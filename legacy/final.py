import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path.cwd()

INPUT = ROOT / "raw_text" / "word_documents.csv"
OUTPUT = ROOT / "Chaput_Document_Archive.xlsx"


# ============================================================
# LOAD CATALOG
# ============================================================

with INPUT.open(
    encoding="utf-8-sig",
    newline=""
) as f:
    records = list(csv.DictReader(f))


# ============================================================
# SORT BY WORD COUNT
# ============================================================

records.sort(
    key=lambda r: int(r["word_count"]),
    reverse=True
)


# ============================================================
# CREATE WORKBOOK
# ============================================================

wb = Workbook()

ws = wb.active
ws.title = "Document Catalog"


headers = [
    "Rank",
    "Word Count",
    "Filename",
    "Location",
    "Modified",
    "Raw Text",
    "Preview",
]

ws.append(headers)


# ============================================================
# ADD DOCUMENTS
# ============================================================

for rank, record in enumerate(records, 1):

    file_location = record["path"].replace("\\", "/")

    raw_text = record["text_file"].replace("\\", "/").replace(
        "_doc_catalog/",
        "raw_text/",
        1,
    )

    filename = record["filename"]

    ws.append([
        rank,
        int(record["word_count"]),
        filename,
        file_location,
        record["modified"],
        raw_text,
        record["preview"],
    ])

    row = ws.max_row

    # --------------------------------------------------------
    # Filename → original Word document
    # --------------------------------------------------------

    filename_cell = ws.cell(row, 3)

    filename_cell.hyperlink = (
        f"./{file_location}"
    )

    filename_cell.font = Font(
        color="0563C1",
        underline="single",
    )

    # --------------------------------------------------------
    # File Location → original Word document
    # --------------------------------------------------------

    location_cell = ws.cell(row, 4)

    location_cell.hyperlink = (
        f"./{file_location}"
    )

    location_cell.font = Font(
        color="0563C1",
        underline="single",
    )

    # --------------------------------------------------------
    # Raw Text → extracted text file
    # --------------------------------------------------------

    raw_cell = ws.cell(row, 6)

    raw_cell.hyperlink = (
        f"./{raw_text}"
    )

    raw_cell.font = Font(
        color="0563C1",
        underline="single",
    )


# ============================================================
# FORMAT
# ============================================================

# Freeze header row
ws.freeze_panes = "A2"

# Enable filtering
ws.auto_filter.ref = ws.dimensions


# Header
for cell in ws[1]:

    cell.font = Font(bold=True)

    cell.alignment = Alignment(
        vertical="center"
    )


# Column widths
ws.column_dimensions["A"].width = 7
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 42
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 11
ws.column_dimensions["G"].width = 80


# Word count formatting
for row in ws.iter_rows(min_row=2):

    row[1].number_format = '#,##0'

    row[6].alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )


# ============================================================
# EXCEL TABLE (formatting rejected by Excel, so I commented it out)
# ============================================================

# table = Table(
#     displayName="DocumentCatalog",
#     ref=f"A1:G{ws.max_row}",
# )

# style = TableStyleInfo(
#     name="TableStyleMedium2",
#     showFirstColumn=False,
#     showLastColumn=False,
#     showRowStripes=True,
#     showColumnStripes=False,
# )

# table.tableStyleInfo = style

# ws.add_table(table)

# ws.row_dimensions[1].height = 24


# ============================================================
# SAVE
# ============================================================

wb.save(OUTPUT)

print()
print("=" * 60)
print("CATALOG CREATED")
print("=" * 60)
print()
print(f"Documents: {len(records):,}")
print(f"Output:    {OUTPUT}")
print()
print("Sorted by word count: largest → smallest")
print("Clickable: Filename, File Location, Raw Text")
print("Header row: frozen")