# images_final.py

import csv
import re
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

try:
    from PIL import Image
except ImportError:
    Image = None


ROOT = Path.cwd()
OUTPUT = ROOT / "Chaput_Image_Archive.xlsx"

IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif",
    ".bmp", ".tif", ".tiff"
}


def clean_excel_text(value):
    if not value:
        return ""

    return re.sub(
        r"[\x00-\x08\x0B\x0C\x0E-\x1F]",
        "",
        str(value),
    )


def image_type(path):
    return path.suffix.lower().lstrip(".").upper()


def get_preview(path):
    if Image is None:
        return ""

    try:
        with Image.open(path) as img:
            width, height = img.size
            return f"{width:,} × {height:,}"
    except Exception:
        return ""


# ============================================================
# FIND IMAGES
# ============================================================

files = [
    p for p in ROOT.rglob("*")
    if p.is_file()
    and p.suffix.lower() in IMAGE_EXTENSIONS
    and not p.name.startswith("~$")
    and "Chaput_Image_Catalog.xlsx" not in str(p)
]

print()
print("=" * 60)
print("CHAPUT IMAGE CATALOG")
print("=" * 60)
print()
print(f"Found {len(files):,} images")
print()


# ============================================================
# COLLECT METADATA
# ============================================================

records = []

for i, path in enumerate(files, 1):

    relative = path.relative_to(ROOT)
    location = str(relative).replace("\\", "/")

    try:
        modified = datetime.fromtimestamp(
            path.stat().st_mtime
        ).isoformat(
            sep=" ",
            timespec="seconds",
        )
    except Exception:
        modified = ""

    records.append({
        "type": image_type(path),
        "filename": path.name,
        "location": location,
        "modified": modified,
        "preview": get_preview(path),
    })

    if i % 100 == 0 or i == len(files):
        print(f"Processed {i:,}/{len(files):,}")


# ============================================================
# SORT
# ============================================================

records.sort(
    key=lambda r: (
        r["type"].lower(),
        r["filename"].lower(),
    )
)


# ============================================================
# CREATE WORKBOOK
# ============================================================

wb = Workbook()
ws = wb.active
ws.title = "Image Catalog"

headers = [
    "Type",
    "Filename",
    "Location",
    "Modified",
    "Dimensions",
]

ws.append(headers)


for record in records:

    ws.append([
        record["type"],
        record["filename"],
        record["location"],
        record["modified"],
        record["preview"],
    ])

    row = ws.max_row

    # Filename hyperlink
    filename_cell = ws.cell(row, 2)
    filename_cell.hyperlink = (
        f"./{record['location']}"
    )
    filename_cell.font = Font(
        color="0563C1",
        underline="single",
    )

    # Location hyperlink
    location_cell = ws.cell(row, 3)
    location_cell.hyperlink = (
        f"./{record['location']}"
    )
    location_cell.font = Font(
        color="0563C1",
        underline="single",
    )

    # Preview
    ws.cell(row, 5).alignment = Alignment(
        vertical="top"
    )


# ============================================================
# FORMAT
# ============================================================

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(
        vertical="center"
    )

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 65
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 18

ws.row_dimensions[1].height = 24


# ============================================================
# SAVE
# ============================================================

wb.save(OUTPUT)

print()
print("=" * 60)
print("CATALOG CREATED")
print("=" * 60)
print()
print(f"Images: {len(records):,}")
print(f"Output: {OUTPUT}")
print()
print("Sorted by type → filename")
print("Clickable: Filename, Location")
print("Preview: image dimensions")
print("Header row: frozen")