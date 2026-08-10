from pathlib import Path
from datetime import datetime
import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from mutagen import File as MutagenFile
except ImportError:
    print("Install mutagen first:")
    print("pip install mutagen")
    sys.exit(1)


# ============================================================
# CONFIGURATION
# ============================================================

ROOT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()

OUTPUT = ROOT / "Chaput_Audio_Archive.xlsx"

AUDIO_EXTENSIONS = {
    ".mp3",
    ".wma",
    ".m4a",
    ".wav",
    ".ogg",
    ".m3u",
    ".wpl",
}

# ============================================================
# HELPERS
# ============================================================

def format_length(seconds):
    if seconds is None:
        return ""

    try:
        seconds = int(round(seconds))
    except (TypeError, ValueError):
        return ""

    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"

    return f"{minutes}:{seconds:02d}"


def get_length(path):
    try:
        audio = MutagenFile(path)

        if audio is not None and audio.info:
            return format_length(
                getattr(audio.info, "length", None)
            )

    except Exception:
        pass

    return ""


def clean_excel_text(value):
    if value is None:
        return ""

    # Remove characters Excel/openpyxl cannot store.
    return "".join(
        char
        for char in str(value)
        if ord(char) in (9, 10, 13)
        or ord(char) >= 32
    )


# ============================================================
# FIND AUDIO
# ============================================================

audio_files = [
    p for p in ROOT.rglob("*")
    if p.is_file()
    and p.suffix.lower() in AUDIO_EXTENSIONS
    and not any(
        part.startswith("_")
        for part in p.relative_to(ROOT).parts
    )
]

print()
print("=" * 60)
print("CHAPUT AUDIO CATALOG")
print("=" * 60)
print()
print(f"Found {len(audio_files):,} audio files")
print()


records = []

# ============================================================
# PROCESS
# ============================================================

for i, path in enumerate(audio_files, 1):

    relative = path.relative_to(ROOT)

    location = str(relative.parent)

    if location == ".":
        location = ""

    filename = path.name

    try:
        modified = datetime.fromtimestamp(
            path.stat().st_mtime
        ).isoformat(
            sep=" ",
            timespec="seconds",
        )
    except Exception:
        modified = ""

    length = get_length(path)

    records.append({
        "type": path.suffix.lower().lstrip(".").upper(),
        "filename": filename,
        "location": location,
        "modified": modified,
        "length": length,
        "_sort_location": location.lower(),
        "_sort_filename": filename.lower(),
    })

    if i % 25 == 0 or i == len(audio_files):
        print(
            f"Processed {i:,}/{len(audio_files):,}"
        )


# ============================================================
# SORT
# ============================================================

# records.sort(
#     key=lambda r: (
#         r["_sort_location"],
#         r["_sort_filename"],
#     )
# )

records.sort(
    key=lambda r: r["_sort_filename"]
)

records.sort(
    key=lambda r: r["_sort_location"],
    reverse=True
)


# ============================================================
# CREATE WORKBOOK
# ============================================================

wb = Workbook()

ws = wb.active
ws.title = "Audio Catalog"


headers = [
    "Type",
    "Filename",
    "Location",
    "Modified",
    "Length",
]

ws.append(headers)


# ============================================================
# ADD RECORDS
# ============================================================

for record in records:

    filename = record["filename"]

    relative_path = Path(
        record["location"]
    ) / filename

    relative_path = str(
        relative_path
    ).replace("\\", "/")

    ws.append([
        record["type"],
        filename,
        record["location"],
        record["modified"],
        record["length"],
    ])

    row = ws.max_row

    # --------------------------------------------------------
    # Filename → audio file
    # --------------------------------------------------------

    filename_cell = ws.cell(row, 2)

    filename_cell.hyperlink = (
        f"./{relative_path}"
    )

    filename_cell.font = Font(
        color="0563C1",
        underline="single",
    )

    # --------------------------------------------------------
    # Location → containing folder
    # --------------------------------------------------------

    location_cell = ws.cell(row, 3)

    if record["location"]:

        location_path = (
            record["location"]
            .replace("\\", "/")
        )

        location_cell.hyperlink = (
            f"./{location_path}"
        )

        location_cell.font = Font(
            color="0563C1",
            underline="single",
        )


# ============================================================
# FORMAT
# ============================================================

ws.freeze_panes = "A2"

ws.auto_filter.ref = ws.dimensions


# Header
for cell in ws[1]:

    cell.font = Font(bold=True)

    cell.alignment = Alignment(
        vertical="center"
    )


# Column widths
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 65
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 12


# Vertical alignment
for row in ws.iter_rows(min_row=2):

    for cell in row:

        cell.alignment = Alignment(
            vertical="top"
        )


ws.row_dimensions[1].height = 24


# ============================================================
# EXCEL TABLE (formatting rejected by Excel, so I commented it out)
# ============================================================

# table = Table(
#     displayName="AudioCatalog",
#     ref=f"A1:E{ws.max_row}",
# )

# style = TableStyleInfo(
#     name="TableStyleMedium2",
#     showFirstColumn=False,
#     showLastColumn=False,
#     showRowStripes=True,
#     showColumnStripes=False,
# )

# table.tableStyleInfo = style

# ws.add_table(table)


# ============================================================
# SAVE
# ============================================================

wb.save(OUTPUT)

print()
print("=" * 60)
print("CATALOG CREATED")
print("=" * 60)
print()
print(f"Audio files: {len(records):,}")
print(f"Output:      {OUTPUT}")
print()
print("Sorted by: Location → Filename")
print("Clickable: Filename, Location")
print("Length:    detected from audio metadata")
print("Header:    frozen")